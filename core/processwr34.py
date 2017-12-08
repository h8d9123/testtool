#-*-coding:utf-8-*-
from PyQt4 import QtGui,QtCore,Qt
import openpyxl as pyxl
import numpy as np
import os

CHANNEL_COUNT = 8
CHANNEL_1 = 3
IL = 0
ILVAR = 1
ILRIP = 2
ILRIPR = 3
GDRIP1 = 4
GDRIPR1 = 5
GDRIP2 = 6
GDRIPR2 = 7
RJ = 8
RL = 9
CPRL=10
COLSPEC =   2
COLT1   =   3
COLT2   =   4
COLT3   =   5
COLPASSORFAIL = 6

def parseSpecStr(spec):
    '''
    Desp:����ָ��,�жϴ���С�ڣ��Լ�ָ�꣬���Ե�λ
    Args:
        spec: �ַ����� <=/>= + ָ�� + ��λ�� ���� ����0.15dB/MHz
        ���أ�
            ����ָ��ֵ
        ���⣺
        <=��>= δ�����������й�
    '''
    endIdx = 1
    for i in range(1,len(spec)):
        if str(spec[1]) == '+' or str(spec[1]) == '-':
            continue
        if spec[i].isdigit() or str(spec[i]) == '.':
            continue
        else:
            endIdx = i
            break
    #isSmall = 1 if ord(spec[0])==ord('��') else 0
    if len(spec[1:endIdx])==0:
        return 0
    v = float(str(spec[1:endIdx]))
    return v
    

class WR34Filter(object):
    def __init__(self,stdExcellName,sheetName=None):
        self.chanelCount = CHANNEL_COUNT    #ͨ������
        self.rowStep = 11       #ÿ��ͨ����Ӧ����������
        self.stdExcelName = stdExcellName   #ָ��ģ��-->output_wr34.xlsx
        self.wb = pyxl.load_workbook(self.stdExcelName)
        #��ȡ������
        if sheetName is None:
            self.ws = self.wb.active
        else:
            if sheetName in self.wb.get_sheet_names():
                self.ws = self.wb.get_sheet_by_name(sheetName)
            else:
                self.ws = self.active
        #ÿһ��ͨ����Ҫд���ָ��λ�ã� ����Ӧ��������
        self.specName = [IL,ILRIP,ILRIPR,GDRIP1,GDRIPR1,GDRIP2,GDRIPR2,RJ,RL,CPRL]
        #�¶ȶ�Ӧ����
        self.colTemparaturs = [COLT1, COLT2, COLT3]
        #��¼ÿ��ͨ���ĸ���ָ�꣬�����С�� ͨ���� x ÿ��ͨ�����ݵĸ���
        self.specLim = np.zeros(shape = (self.chanelCount, self.rowStep))
        #��¼��>= ����<=
        self.isSmallEqual = np.ones(self.rowStep)
        self.isSmallEqual[-3:]=0
        
        self.readSpecLimition()
    def readSpecLimition(self):
        '''
        Desp:
                        ��ģ���ж�ȡÿһ��ָ���Ӧ�ķ�Χ
        '''
        for ch in range(self.chanelCount):
            for r in range(self.rowStep):
                #�� ch ��ͨ���� ÿһ��ͨ������������Ϊ rowstep, ��ʼ��CHANEL_1, r:ÿ��ͨ���ĵ�r��������
                row_excel = ch*self.rowStep + CHANNEL_1 + r
                spec = self.ws[row_excel][COLSPEC].value
                v = parseSpecStr(spec)
                self.specLim[ch, r] = v
        #print self.specLim
    def writeSpecToExcel(self, targetExcel, chanel,colTemperature, specs):
        wb = None
        ws = None
        
        if os.path.exists(targetExcel): #���excel����,��д��Ŀ���ļ�����������ģ���ļ����Ϊ�µ�Ŀ���ļ�
            wb = pyxl.load_workbook(targetExcel)
            ws = wb.active
        else:
            wb = self.wb
            ws = self.ws
        #д���s2p�õ��� ͨ��chanel, ���¶ȶ�Ӧ������
        specName = self.specName
        for idxName in specName:
            #�� chanel ��ͨ���� ÿһ��ͨ������������Ϊ rowstep, ��ʼ��CHANEL_1, r:ÿ��ͨ���ĵ�r��������
            row_excel = (chanel-1)*self.rowStep + CHANNEL_1 + idxName
            ws[row_excel][colTemperature].value = specs[idxName]
        vec =[]
        row_excel = (chanel-1)*self.rowStep + CHANNEL_1
        for col in self.colTemparaturs:
            tmp = ws[row_excel][col].value
            if tmp!='' and tmp is not None:
                print [tmp]
                vec.append(float(tmp))
            else:
                vec.append(0)
        ws[row_excel + 1][self.colTemparaturs[0]].value = np.max(vec) - np.min(vec)
        
        
        #�жϵ�ǰͨ���������Ƿ���
        self.checkPassOrFail(ws, chanel)
        #���excel�ļ��Ѿ������쳣����ʾ�ر��ٱ���
        try:
            wb.save(targetExcel)
        except IOError,e:
            if e.args[0]==13:
                QtGui.QMessageBox.warning(None,"error","Please Close %s, firstly"%(e.filename))
            return False
        
        return True
    def checkPassOrFail(self, ws,chanel):
        specName = self.specName
        for idxName in specName:
            #�� chanel ��ͨ���� ÿһ��ͨ������������Ϊ rowstep, ��ʼ��CHANEL_1, r:ÿ��ͨ���ĵ�r��������
            row_excel = (chanel-1)*self.rowStep + CHANNEL_1 + idxName
            #�ӱ���ȡͬһͨ����ͬ�¶��µ�ʵ��ֵ����Ϊ����Ϊ -1
            vec = []
            for col in self.colTemparaturs:
                v = ws[row_excel][col].value
                v = float(v) if v !='' and  v!=None else -1
                vec.append(v)
            if self.isSmallEqual[idxName]==1:
                ws[row_excel][COLPASSORFAIL].value = 'PASS' if np.max(vec)<=self.specLim[chanel - 1, idxName] and np.min(vec)!=-1 else 'FAIL'
            else:
                ws[row_excel][COLPASSORFAIL].value = 'PASS' if np.min(vec)>=self.specLim[chanel - 1, idxName] and np.min(vec)!=-1 else 'FAIL'
        row_excel = (chanel-1)*self.rowStep + CHANNEL_1
        t = ws[row_excel + 1][self.colTemparaturs[0]].value
        if t!='' and t!=None:
            if float(t) <=self.specLim[chanel -1, ILVAR]:
                ws[row_excel+1][COLPASSORFAIL].value = 'PASS'
            else:
                ws[row_excel+1][COLPASSORFAIL].value = 'Fail'
        else:
            ws[row_excel+1][COLPASSORFAIL].value = 'Fail'
        pass

def getInputFrequency(excelName):
    def parseStr(s):
        endIdx = 0
        for i in range(0,len(s)):
            if s[i].isdigit() or str([i]) == '.':
                continue
            else:
                endIdx = i
                break
        if len(s[0:endIdx])==0:
            return 0
        
        #Ĭ��MHz��תΪHz
        v = float(str(s[0:endIdx]))*10**6
        return v
    wb = pyxl.load_workbook(excelName)
    ws = wb.active
    vec_cf_bw = np.zeros(shape = (CHANNEL_COUNT, 3))
    starRow = 2
    for r in range(CHANNEL_COUNT):
        v1 = ws[r + starRow][2].value
        vec_cf_bw[r,0] =parseStr(v1)
        v2 = ws[r+starRow][3].value
        vec_cf_bw[r,1] = parseStr(v2)
        v3 = ws[r+starRow][4].value
        vec_cf_bw[r,2] = parseStr(v3)
    #print vec_cf_bw
    return vec_cf_bw

def getchAndJIndex(excelName):
    wb = pyxl.load_workbook(excelName)
    ws = wb.active
    vec_cf_bw = np.zeros(shape = (CHANNEL_COUNT, 2))
    starRow = 2
    v=[1 for i in range(CHANNEL_COUNT)]
    for r in range(CHANNEL_COUNT):
        v1 = ws[r + starRow][1].value
        v[r]=int(v1[1:])
    return v
        
        
        
if __name__ == '__main__':
    import sys
    app = QtGui.QApplication(sys.argv)
    stdExcel = r'..\db\output_WR34.xlsx'
    tarExcel = r'..\qa\outputDir\out1.xlsx'
    inputExcel = r'..\db\input_WR34.xlsx'
    wr34 = WR34Filter(stdExcel)
    specs = np.zeros(11);
    ch = 2
    colT =COLT1
    #wr34.writeSpecToExcel(tarExcel, ch, colT, specs)
    vecs = getInputFrequency(inputExcel)
    vv = getchAndJIndex(inputExcel)
    print vv.index(8)
    sys.exit(app.exec_())
    pass